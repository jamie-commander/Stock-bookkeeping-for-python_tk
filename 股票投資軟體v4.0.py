import requests
import urllib.request
import os
import re
import random
import json
import time
import sys
from tkinter import *
from tkinter import ttk
from lxml import html
from threading import Timer
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import colors, Font, Fill, NamedStyle
from openpyxl.styles import PatternFill, Border, Side, Alignment
import Catch_BigData as Stock_Api
root = Tk()
root.title("股票投資軟體v4")
root.geometry("1650x825+0+0")
global filepath
filepath = os.getcwd()
global file
file="data"
global myfile
myfile=list()
'''try:
    wb = load_workbook('股票倉庫與紀錄.xlsx')
except:
    wb = Workbook()
    wb[wb.sheetnames[0]].title='股票交易紀錄'#第n工作表名稱
    wb.create_sheet(index=1, title='股票倉庫')
ws1 = wb[wb.sheetnames[0]]
ws2 = wb[wb.sheetnames[1]]'''
data=list()
try:
    with open(os.path.join(filepath,"remanber.txt"),'r') as f:
        file=f.read()
except:
    with open("remanber.txt","w") as f:
        f.write("data")
        file="data"
try:
    with open(os.path.join(filepath,file+".json"),'r') as f:
        data =json.load(f)
except:
    #print("data尚未擁有任何資料")
    with open(file+".json","w") as f:
        json.dump(data,f)
for i in os.listdir(filepath):
    if(i[len(i)-5:len(i)]==".json"):
        #print(i[0:len(i)-5])
        myfile.append(i[0:len(i)-5])
#print(data,tdata)
table=list()
dtable=list()

taba=list()
tabb=list()
gtable=list()
global page
page=1
global cot
cot=0
global pagen
pagen=1
global pp
pp=0
global ppa
ppa=0

#wb.save('股票倉庫與紀錄.xlsx')
def updata():
    t=Timer(0,ck,[])
    t.start()
def ck():
    ha.set("資料更新中，請稍後。")
    Api = Stock_Api.Catch_Stocks_BigData()
    Api.ALL_Data_main()
    tse_data = dict()
    otc_data = dict()
    filepath = os.path.join(os.getcwd(),"BigData")
    filepath_tse = os.path.join(filepath,"tse")
    filepath_otc = os.path.join(filepath,"otc")
    path_tse = os.path.join(filepath_tse,"Stocks")
    path_otc = os.path.join(filepath_otc,"Stocks")
    path_list = os.listdir(path_tse)
    if(len(path_list) == 0):
        date = date
    else:
        for i in range(0,len(path_list)):
            path_list[i] = int(path_list[i][0:(len(path_list[i]) - 5)])
        date =str(max(path_list))
    with open(os.path.join(path_tse, date +".json"),'r') as f:
        tse_data =json.load(f)
    #print(tse_data["title_name"])
    path_list = os.listdir(path_otc)
    if(len(path_list) == 0):
        date = date
    else:
        for i in range(0,len(path_list)):
            path_list[i] = int(path_list[i][0:(len(path_list[i]) - 5)])
        date =str(max(path_list))
    with open(os.path.join(path_otc, date +".json"),'r') as f:
        otc_data =json.load(f)
    #print(tse_data["Stocks_data"]["03014X"])
    #print(tse_data["title_name"])
    global cot
    if(cot==1):
        return
    global data
    for i in range(0,len(data)):
        if(str(data[i][16])=="tyo"):
            url="https://mis.tpex.org.tw/Quote.asmx/GETQ20?SymbolID="
            url=url+str(data[i][0])
            while True:
                try:
                    gs = requests.get(url,timeout=3)
                    dom = BeautifulSoup(gs.text, 'html.parser')
                    go=dom.find("tradeprice")#興櫃成交價
                    ti=dom.find("tradestatistictime")
                    data[i][3]=float(go.text)
                    break
                except:
                    time.sleep((random.randint(1,2))/10)
            try:
                if(0<len(data)):
                    ga.set("上次更新時間："+ti.text)
                else:
                    ga.set("上次更新時間："+"請載入資料")
            except:
                ga.set(ga.get())
                
        else:
            try:
                if(str(data[i][16])=="tse"):
                    data[i][3] = float(tse_data["Stocks_data"][str(data[i][0])][8])
                if(str(data[i][16])=="otc"):
                    data[i][3] = float(otc_data["Stocks_data"][str(data[i][0])][8])
                if(0<len(data)):
                    ga.set("上次更新時間："+str(st["queryTime"]["sysTime"]))
                else:
                    ga.set("上次更新時間："+"請載入資料")
            except:
                ga.set(ga.get())
    mo=0;
    for i in range(0,12):
        put(i)
    ha.set("資料更新完成。")
    return
def put(kd):
    global pagen,page,pp,ppa,file
    for i in range(0,len(data)):
        exc(data[i])
    if(pp==0):
        to=kd+(page-1)*12
        pa.set(str(page)+"/"+str(int((len(data)-1)/12)+1))
        if(to<len(data)):
            for i in range(0,14):
                if(i<2):
                    table[kd][i].set(str(data[to][i]))
                elif(i==7):
                    if(float(data[to][i]<0)):
                        taba[kd*25+38].config(bg="#FAF0E6",fg="#22AA22",font=('微軟正黑體',11,'bold'),relief=GROOVE)
                    else:
                        taba[kd*25+38].config(bg="#FAF0E6",fg="#AA2222",font=('微軟正黑體',11,'bold'),relief=GROOVE)
                    table[kd][i].set(str(int(data[to][i])))
                elif(i==2 or i==4 or i==5 or i==6 or i==7 or i==8 or i==9 or i==10):
                    table[kd][i].set(str(int(data[to][i])))
                elif(i==11):
                    if(int(data[to][2])==0):
                        table[kd][i].set(str("清空"))
                    else:
                        table[kd][i].set(str(round(float(data[to][i]),2)))
                elif(i<12):
                    table[kd][i].set(str(round(float(data[to][i]),2)))
                elif(i<14):
                    if(float(data[to][i]<0)):
                        taba[kd*25+44].config(bg="#FAF0E6",fg="#22AA22",font=('微軟正黑體',11,'bold'),relief=GROOVE)
                        taba[kd*25+43].config(bg="#FAF0E6",fg="#22AA22",font=('微軟正黑體',11,'bold'),relief=GROOVE)
                    else:
                        taba[kd*25+44].config(bg="#FAF0E6",fg="#AA2222",font=('微軟正黑體',11,'bold'),relief=GROOVE)
                        taba[kd*25+43].config(bg="#FAF0E6",fg="#AA2222",font=('微軟正黑體',11,'bold'),relief=GROOVE)
                    if(i==12):
                        table[kd][i].set(str(int(data[to][i])))
                    else:
                        table[kd][i].set(str(round(float(data[to][i]),2)))
            table[kd][17].set(data[to][14])
            table[kd][18].set(data[to][15])
        else:
            for i in range(0,17):
                table[kd][i].set("")
        #total=["投資成本","賣出成本","賣出匯入","買賣報酬率","手續費","股票市值","買賣損益","自訂義損益","報酬率"]
        #toy=["買入股數","股票股利","現金股利","平均成交價","賣出股數","平均賣出價","手續費","買賣報酬率","股票市值"]
    else:
        to=kd+(pagen-1)*12
        pa.set(str(pagen)+"/"+str(int((len(data[ppa][17])-1)/12)+1))
        if(to<len(data[ppa][17])):
            for i in range(0,11):
                if(i==4): 
                    gtable[kd][i].set(int(data[ppa][17][to][i]))
                else:
                    gtable[kd][i].set(data[ppa][17][to][i])
        else:
            for i in range(0,11):
                gtable[kd][i].set("")
    for k in range(4,11):
        co=0
        for j in range(0,len(data)):
            co=co+data[j][k]
        dtable[k-4].set(str(int(co)))
        if(k==7):
            if(co<0):
                tabc[3].config(bg="#FAF0E6",fg="#22AA22",font=('微軟正黑體',16,'bold'),relief=GROOVE)
            else:
                tabc[3].config(bg="#FAF0E6",fg="#AA2222",font=('微軟正黑體',16,'bold'),relief=GROOVE)
    co=0
    for j in range(0,len(data)):
        co=co+data[j][12]
    if(co<0):
        tabc[7].config(bg="#FAF0E6",fg="#22AA22",font=('微軟正黑體',16,'bold'),relief=GROOVE)
        tabc[8].config(bg="#FAF0E6",fg="#22AA22",font=('微軟正黑體',16,'bold'),relief=GROOVE)
    else:
        tabc[7].config(bg="#FAF0E6",fg="#AA2222",font=('微軟正黑體',16,'bold'),relief=GROOVE)
        tabc[8].config(bg="#FAF0E6",fg="#AA2222",font=('微軟正黑體',16,'bold'),relief=GROOVE)
    dtable[7].set(str(int(co)))
    if(int(dtable[1].get())!=0):
        co=int(dtable[7].get())/int(dtable[1].get())*100
    else:
        co=0
    dtable[8].set(str(round(co,3)))
    with open(file+".json","w") as f:
        json.dump(data,f)
    with open("remanber.txt","w") as f:
        f.write(file)
    return
def pagedown():
    global page,pagen,pp,ppa
    if(pp==0):
        if(page<int((len(data)-1)/12)+1):
            page=page+1
        pa.set(str(page)+"/"+str(int((len(data)-1)/12)+1))
    else:
        if(pagen<int((len(data[ppa][17])-1)/12)+1):
            pagen=pagen+1
        pa.set(str(pagen)+"/"+str(int((len(data[ppa][17])-1)/12)+1))
    for i in range(0,12):
        put(i)
    return
def pageup():
    global page,pagen,pp,ppa
    if(pp==0):        
        if(page>1):
            page=page-1
        pa.set(str(page)+"/"+str(int((len(data)-1)/12)+1))
    else:
        if(pagen>1):
            pagen=pagen-1
        pa.set(str(pagen)+"/"+str(int((len(data[ppa][17])-1)/12)+1))
    for i in range(0,12):
        put(i)
    return
def sell(kd):
    ha.set("請輸入正確資料格式")
    global page
    to=kd+(page-1)*12
    tua=table[kd][16].get().split("/")
    if(to>len(data)-1):
        return
    if(table[kd][14].get()=="" or table[kd][15].get()=="" or len(tua)!=3 or data[to][2]+0.1<float(table[kd][15].get())):
        return
    else:
        uu=float(table[kd][14].get())*float(table[kd][15].get())
        if(data[to][14]==0):
            rb=0.0007125
        elif(data[to][14]==1):
            rb=0.001425
        if(data[to][15]==0):
            ra=0.003
        elif(data[to][15]==1):
            ra=0.001
        sa=[data[to][0],data[to][1],"賣出",table[kd][16].get(),float(table[kd][15].get()),float(table[kd][14].get()),int(uu*rb+0.5),int(uu*ra+0.5),0,0,int(uu-uu*(ra+rb)+0.5)]
        exc(data[to],sa)
        data[to][17].append(sa)
        ha.set("賣出"+str(data[to][0])+str(data[to][1])+" 單價:"+str(table[kd][14].get())+"元 股:"+str(table[kd][15].get()))
        rq(data[to][17])
        for i in range(0,10):
            put(i)
        table[kd][14].set("")
        table[kd][15].set("")
        table[kd][16].set("")
    return
def exc(ss,sa=[0,0,0,0,0,0,0,0,0,0,0]):
    if(sa[2]=="買進"):
        ss[2]=ss[2]+sa[4]
        ss[5]=ss[5]+sa[10]
    if(sa[2]=="賣出"):
        ss[2]=ss[2]-sa[4]
        ss[6]=ss[6]+sa[10]
    if(sa[2]=="除權"):
        ss[2]=ss[2]+sa[4]
    if(sa[2]=="除息"):
        ss[8]=ss[8]+sa[9]
    if(sa[2]=="除權息"):
        ss[2]=ss[2]+sa[4]
        ss[8]=ss[8]+sa[9]
    if(sa[2]=="當沖"):
        ss[7]=ss[7]+sa[8]
    ss[4]=ss[2]*ss[3]
    if(ss[14]==0):
        ss[9]=int(ss[4]*0.0007125+0.5)
    elif(ss[14]==1):
        ss[9]=int(ss[4]*0.001425+0.5)
    if(ss[15]==0):
        ss[10]=int(ss[4]*0.003+0.5)
    elif(ss[15]==1):
        ss[10]=int(ss[4]*0.001+0.5)
    if(ss[2]!=0):
        ss[11]=(ss[5]-ss[6]+ss[7]-ss[8]+ss[9]+ss[10])/ss[2]
    else:
        ss[11]=(ss[5]-ss[6]+ss[7]-ss[8]+ss[9]+ss[10])/0.00001
    ss[12]=ss[4]+ss[6]+ss[7]+ss[8]-ss[5]-ss[9]-ss[10]
    if(ss[5]!=0):
        ss[13]=ss[12]/ss[5]*100
    else:
        ss[13]=0
    return
def buy():
    ha.set("請輸入正確資料格式")
    if(buylist[0].get()==""):
        buylist[0].set("2330")
    if(buylist[1].get()==""):
        buylist[1].set("0.0001")
    if(buylist[2].get()==""):
        buylist[2].set("0")
    tua=buylist[3].get().split("/")
    if(buylist[3].get()==""):
        buylist[3].set("1994/09/05")
    if(buylist[2].get()=="0" or buylist[3].get()=="1994/09/05" or len(tua)!=3):
        return
    kk=0
    con=0
    for i in range(0,len(data)):
        if(buylist[0].get()==data[i][0]):
            kk=1
            con=i
    if(kk==0):        
        ss=list()
        for i in range(0,17):
            ss.append(0)
        kp=list()
        ss.append(kp)
        out_time=0
        while True:
            try:
                out_time = out_time+1
                rs=requests.session()
                url="https://mis.twse.com.tw/stock/api/getStockInfo.jsp?ex_ch=tse_"+buylist[0].get()+".tw"
                rs = requests.get(url,timeout=2)
                #time.sleep((random.randint(1,3))/10)
                ts=requests.session()
                url="https://mis.twse.com.tw/stock/api/getStockInfo.jsp?ex_ch=otc_"+buylist[0].get()+".tw"
                ts = requests.get(url,timeout=2)
                if(len(rs.text)>len(ts.text)):
                    st=json.loads(rs.text[12:len(rs.text)])
                    ss[16]="tse"
                    td=st["msgArray"][0]
                    '''gy=td["b"].split("_")
                    gz=td["a"].split("_")
                    gyb=td["g"].split("_")
                    gza=td["f"].split("_")
                    if(float(gyb[0])>float(gza[0])):
                        ss[3]=float(gz[0])
                    else:
                        ss[3]=float(gy[0])'''
                    if(td["z"]!="-"):
                        ss[3]=float(td["z"])
                    ss[0]=buylist[0].get()#證券編號
                    ss[1]=td["n"]#證券名稱
                elif(len(rs.text)<len(ts.text)):
                    st=json.loads(ts.text[12:len(ts.text)])
                    ss[16]="otc"
                    td=st["msgArray"][0]
                    '''gy=td["b"].split("_")
                    gz=td["a"].split("_")
                    gyb=td["g"].split("_")
                    gza=td["f"].split("_")
                    if(float(gyb[0])>float(gza[0])):
                        ss[3]=float(gz[0])
                    else:
                        ss[3]=float(gy[0])'''
                    if(td["z"]!="-"):
                        ss[3]=float(td["z"])
                    ss[0]=buylist[0].get()#證券編號
                    ss[1]=td["n"]#證券名稱
                else:
                    url="https://mis.tpex.org.tw/Quote.asmx/GETQ20?SymbolID="+buylist[0].get()
                    gs = requests.get(url,timeout=2)
                    dom = BeautifulSoup(gs.text, 'html.parser')
                    gk=dom.find("symbolname")#興櫃名稱
                    go=dom.find("tradeprice")#興櫃成交價
                    ss[16]="tyo"
                    ss[0]=buylist[0].get()
                    ss[1]=gk.text
                    ss[3]=float(go.text)
                break
            except:
                if(out_time>2):
                    ha.set("請勿輸入不存在的股票代碼")
                    return
                time.sleep((random.randint(1,2))/10)
        
        #columns = ['c','n','z','tv','v','o','h','l','y']#columns = ['股票代號','公司簡稱','當盤成交價','當盤成交量','累積成交量','開盤價','最高價','最低價','昨收價']
    else:
        ss=data[con]
    if(buylist[4].get()==0):
        gr=0.0007125
    else:
        gr=0.001425
    uu=float(buylist[1].get())*float(buylist[2].get())
    sa=[ss[0],ss[1],"買進",buylist[3].get(),float(buylist[2].get()),float(buylist[1].get()),int(float(uu*gr+0.5)),0,0,0,int(float(uu)+float(uu*gr)+0.5)]
    exc(ss,sa)
    ha.set("買入"+str(ss[0])+str(ss[1])+" 單價:"+str(buylist[1].get())+"元 股:"+str(float(buylist[2].get())))
    for i in range(0,4):
        buylist[i].set("")
    if(kk==0):
        data.insert(0,ss)
        data[0][17].append(sa)
        rq(data[0][17])
    else:
        data[con]=ss
        data[con][17].append(sa)
        rq(data[con][17])
    del ss
    del sa
    for i in range(0,12):
        put(i)
    return
def rq(sq):
    sw=list()
    for i in range(0,len(sq)):
        sw.append(sq[i][3])
    for i in range(0,len(sw)):
        sw[i]=sw[i].split("/")
    for i in range(0,len(sq)+1):
        for j in range(0,len(sq)-1):
            if(int(sw[j][0])<int(sw[j+1][0])):
                gb=sw[j]
                sw[j]=sw[j+1]
                sw[j+1]=gb
                ga=sq[j]
                sq[j]=sq[j+1]
                sq[j+1]=ga
            elif(int(sw[j][1])<int(sw[j+1][1]) and int(sw[j][0])==int(sw[j+1][0])):
                gb=sw[j]
                sw[j]=sw[j+1]
                sw[j+1]=gb
                ga=sq[j]
                sq[j]=sq[j+1]
                sq[j+1]=ga
            elif(int(sw[j][2])<int(sw[j+1][2]) and int(sw[j][0])==int(sw[j+1][0]) and int(sw[j][1])==int(sw[j+1][1])):
                gb=sw[j]
                sw[j]=sw[j+1]
                sw[j+1]=gb
                ga=sq[j]
                sq[j]=sq[j+1]
                sq[j+1]=ga
    del sw
def hi(a,k,kd=0,ff=-1):
    global page
    to=kd+(page-1)*12
    if(to>len(data)-1):
        return
    k.set(a)
    if(ff==0):
        data[to][14]=a
    if(ff==1):
        data[to][15]=a
    for i in range(0,12):
        put(i)
    return
def ex():
    global cot
    cot=1
    root.quit()
    root.destroy()
    sys.exit()
    return
def gonext(kd):
    global page,pagen,pp,ppa
    pp=1
    ppa=kd+(page-1)*12
    if(ppa>len(data)-1):
        return
    toy=["即時市值","投資成本","賣出匯入","當沖","現金股利","手續費","交易稅","帳面損益","損益評估"]
    for i in range(0,9):
        dta[i].set(toy[i])
    for i in range(0,len(taba)):
        taba[i].grid_remove()
    for i in range(0,len(tabb)):
        tabb[i].grid()
    for i in range(0,12):
        put(i)
    return
def goback():
    total=["即時市值","投資成本","賣出匯入","當沖","現金股利","手續費","交易稅","帳面損益","損益評估"]
    global page,pagen,pp,ppa
    pagen=1
    pp=0
    for i in range(0,9):
        dta[i].set(total[i])
    for i in range(0,len(tabb)):
        tabb[i].grid_remove()
    for i in range(0,len(taba)):
        taba[i].grid()
    for i in range(0,12):
        put(i)
    return
def item():
    global ppa
    tua=inlist[3].get().split("/")
    ha.set("請輸入正確資料格式")
    if(inlist[0].get()=="" or len(tua)!=3):
        return
    if(inlist[1].get()==""):
        inlist[1].set("0")
    if(inlist[2].get()==""):
        inlist[2].set("0")
    if(inlist[1].get()=="0" and inlist[2].get()=="0"):
        return
    rw=float(inlist[1].get())*float(inlist[0].get())
    ry=float(inlist[2].get())*float(inlist[0].get())/10
    if(inlist[1].get()=="0"):
        sa=[data[ppa][0],data[ppa][1],"除權",inlist[3].get(),int(ry),0,0,0,0,int(rw+0.5),0]
    elif(inlist[2].get()=="0"):
        sa=[data[ppa][0],data[ppa][1],"除息",inlist[3].get(),int(ry),0,0,0,0,int(rw+0.5),0]
    else:
        sa=[data[ppa][0],data[ppa][1],"除權息",inlist[3].get(),int(ry),0,0,0,0,int(rw+0.5),0]
    data[ppa][17].append(sa)
    exc(data[ppa],sa)
    ha.set("除權息股數:"+inlist[0].get()+" 現金 :"+inlist[1].get()+" 股票:"+inlist[2].get())
    inlist[0].set("")
    inlist[1].set("")
    inlist[2].set("")
    inlist[3].set("")
    rq(data[ppa][17])
    for i in range(0,12):
        put(i)
    return
def ditem():
    global ppa
    ha.set("請輸入正確資料格式")
    tua=inlist[3].get().split("/")
    if(inlist[4].get()==""):
        inlist[4].set("0")
    if(inlist[4].get()=="0" or len(tua)!=3):
        return
    sa=[data[ppa][0],data[ppa][1],"當沖",inlist[3].get(),0,0,0,0,int(float(inlist[4].get())),0,0]
    data[ppa][17].append(sa)
    exc(data[ppa],sa)
    rq(data[ppa][17])
    ha.set("當沖:"+inlist[4].get())
    inlist[4].set("")
    inlist[3].set("")
    for i in range(0,12):
        put(i)
    return
def ext(ss,sa=[0,0,0,0,0,0,0,0,0,0,0]):
    if(sa[2]=="買進"):
        ss[2]=ss[2]-sa[4]
        ss[5]=ss[5]-sa[10]
    if(sa[2]=="賣出"):
        ss[2]=ss[2]+sa[4]
        ss[6]=ss[6]-sa[10]
    if(sa[2]=="除權"):
        ss[2]=ss[2]-sa[4]
    if(sa[2]=="除息"):
        ss[8]=ss[8]-sa[9]
    if(sa[2]=="除權息"):
        ss[2]=ss[2]-sa[4]
        ss[8]=ss[8]-sa[9]
    if(sa[2]=="當沖"):
        ss[7]=ss[7]-sa[8]
    return
def delitem(kd):
    global pagen,ppa,pp
    to=kd+(pagen-1)*12
    #del data[ppa][8][to]
    if(to>len(data[ppa][17])-1):
        return
    if(data[ppa][17][to][2]=="賣出"):
        ext(data[ppa],data[ppa][17][to])
    if(data[ppa][17][to][2]=="買進"):
        ext(data[ppa],data[ppa][17][to])
    if(data[ppa][17][to][2]=="除息"):
        ext(data[ppa],data[ppa][17][to])
    if(data[ppa][17][to][2]=="除權"):
        ext(data[ppa],data[ppa][17][to])
    if(data[ppa][17][to][2]=="除權息"):
        ext(data[ppa],data[ppa][17][to])
    if(data[ppa][17][to][2]=="當沖"):
        ext(data[ppa],data[ppa][17][to])
    del data[ppa][17][to]
    rq(data[ppa][17])
    if(len(data[ppa][17])==0):
        del data[ppa]
        goback()
    for i in range(0,12):
        put(i)
    return
def callbackFunc(Event):
    global file,filepath,myfile
    #file=combo.get()
    #print(combo.get())
    root.focus()
    return
def openf():
    global file,filepath,myfile,data
    if(combo.get()==""):
        return
    ga.set("")
    file=combo.get()
    data.clear()
    with open(os.path.join(filepath,file+".json"),'r') as f:
        data =json.load(f)
    ha.set("開啟檔案成功，檔名："+file)
    ggg.set(file)
    for i in range(0,12):
        put(i)
    return
def newf():
    global file,filepath,myfile,data
    con=0
    if(combo.get()==""):
        return
    ga.set("")
    for i in myfile:
        if(i==combo.get()):
            con=1
    if(con):
        ha.set("另存檔案失敗，此檔案名稱已經存在。")
        return
    myfile.append(combo.get())
    file = combo.get()
    ha.set("另存檔案成功，檔名："+file)
    ggg.set(file)
    combo["value"]=myfile
    for i in range(0,12):
        put(i)
    return
def renewf():
    global file,filepath,myfile,data
    con=0
    if(combo.get()==""):
        return
    ga.set("")
    for i in myfile:
        if(i==combo.get()):
            con=1
    if(con):
        ha.set("新增檔案失敗，此檔案名稱已經存在。")
        return
    myfile.append(combo.get())
    data.clear()
    file = combo.get()
    ha.set("新增檔案成功，檔名："+file)
    ggg.set(file)
    combo["value"]=myfile
    for i in range(0,12):
        put(i)
    return
def up(kd):
    global page
    to=kd+(page-1)*12
    gg=data[to]
    data[to]=data[to-1]
    data[to-1]=gg
    for i in range(0,12):
        put(i)
    return
def down(kd):
    global page
    to=kd+(page-1)*12
    try:
        gg=data[to]
        data[to]=data[to+1]
        data[to+1]=gg
    except:
        gg=data[to]
        data[to]=data[0]
        data[0]=gg
    for i in range(0,12):
        put(i)
    return
#------------標題-------------------------------------------------
frame1 =Frame(bg="#00FFFF",width=1650,height=100,bd=15,relief=GROOVE)#FLAT SUNKEN RAISED GROOVE RIDGE
frame1.pack_propagate(0)
frame1.grid(row = 0,column = 0)
lab=Label(frame1,text="股  票  倉  庫",width=20,height=2,bg="#0000A0",fg="#00FFFF",font=('微軟正黑體',40),relief=GROOVE)
lab.pack(expand=1, fill=BOTH)
#----------------------------買賣欄位------------------------------------------
buy_str=["證券號碼：","買入單價：","買入股數：","日期","手續費"]
buylist=list()
buya=[8,6,12,10]
frame2 =Frame(bg="#C0C0C0",width=1650,height=100,bd=10,relief=GROOVE)
frame2.grid_propagate(0)
frame2.grid(row = 1,column = 0)
lab=Label(frame2,text="手續費：0.1425%，電子交易=0.07125%。交易稅：0.3%，若是ETF=0.1%。",width=116,height=1,bg="#0000A0",fg="#00FFFF",font=('微軟正黑體',18),anchor=W,relief=GROOVE)#EWSN
lab.grid(row = 0,column = 0,columnspan=14,padx=0,pady=0)
for i in range(0,6):
    if(i<4):
        lab=Label(frame2,text=buy_str[i],width=len(buy_str[i])*2,height=1,bg="#0000A0",fg="#00FFFF",font=('微軟正黑體',18),relief=GROOVE)
        taba.append(lab)
        lab.grid(row = 1,column = i*2,padx=2,pady=5)
        buylist.append(StringVar())
        entry = Entry(frame2,textvariable = buylist[i],width=buya[i],font=('微軟正黑體',18),relief=GROOVE)
        entry.grid(row = 1,column = 1+i*2,padx=2,pady=5)
        taba.append(entry)
    elif(i==5):
        lab=Label(frame2,text=buy_str[4],width=len(buy_str[4])*2,height=1,bg="#0000A0",fg="#00FFFF",font=('微軟正黑體',18),relief=GROOVE)
        lab.grid(row = 1,column = 10,padx=2,pady=5)
        taba.append(lab)
        buylist.append(IntVar())
        buylist[4].set(0)
        R=Radiobutton(frame2,variable = buylist[4] ,text ="0.07125%",value = 0,command=lambda k=buylist[4],a=0: hi(a,k),width=8,height=1,bg="#FFFFCD",fg="#292421",font=('微軟正黑體',14),relief=GROOVE)
        R.grid(row=1,column=11,padx=2,pady=5,rowspan=2)
        taba.append(R)
        R=Radiobutton(frame2,variable = buylist[4] ,text ="0.1425%",value = 1,command=lambda k=buylist[4],a=1: hi(a,k),width=8,height=1,bg="#FFFFCD",fg="#292421",font=('微軟正黑體',14),relief=GROOVE)
        R.grid(row=1,column=12,padx=2,pady=5,rowspan=2)
        taba.append(R)
        k=Button(frame2,text= "買入匯出",command=lambda : buy(),width=10, height=1,bg="#004040",fg="#FFFFFF",font=('微軟正黑體',12),relief=GROOVE)
        k.grid(row = 1,column = 13,padx=2,pady=5)
        taba.append(k)
#----------------------------------------------------股票明細------------------------------------
in_str=["除權息股數","每股現金股利","每股股票股利","日期","當沖"]
inlist=list()
inleng=[10,6,6,10]
for i in range(0,4):
    lab=Label(frame2,text=in_str[i],width=len(in_str[i])*2,height=1,bg="#0000A0",fg="#00FFFF",font=('微軟正黑體',16),relief=GROOVE)
    tabb.append(lab)
    lab.grid(row = 1,column = i*2,padx=2,pady=5)
    inlist.append(StringVar())
    entry = Entry(frame2,textvariable = inlist[i],width=inleng[i],font=('微軟正黑體',16),relief=GROOVE)
    entry.grid(row = 1,column = 1+i*2,padx=2,pady=5)
    tabb.append(entry)
k=Button(frame2,text= "除權息",command=lambda : item(),width=10, height=1,bg="#004040",fg="#FFFFFF",font=('微軟正黑體',12),relief=GROOVE)
k.grid(row = 1,column = 9,padx=2,pady=5)
tabb.append(k)

lab=Label(frame2,text=in_str[4],width=len(in_str[4])*2,height=1,bg="#0000A0",fg="#00FFFF",font=('微軟正黑體',16),relief=GROOVE)
tabb.append(lab)
lab.grid(row = 1,column = 10,padx=2,pady=5)
inlist.append(StringVar())
entry = Entry(frame2,textvariable = inlist[4],width=8,font=('微軟正黑體',16),relief=GROOVE)
entry.grid(row = 1,column = 11,padx=2,pady=5)
tabb.append(entry)

k=Button(frame2,text= "當沖",command=lambda : ditem(),width=8, height=1,bg="#004040",fg="#FFFFFF",font=('微軟正黑體',12),relief=GROOVE)
k.grid(row = 1,column = 12,padx=2,pady=5)
tabb.append(k)

k=Button(frame2,text= "回主頁",command=lambda : goback(),width=10, height=1,bg="#004040",fg="#FFFFFF",font=('微軟正黑體',12),relief=GROOVE)
k.grid(row = 1,column = 13,padx=2,pady=5)
tabb.append(k)
#---------------------------------倉庫列表---------------------------
WH_str=["證券碼","證券名稱","庫存數量","即時價","即時市值","投資成本","賣出匯入","當沖","現金股利","手續費","交易稅","均價","即時損益","損益率","賣出價","賣出數","日期","手續費","交易稅"]
ttd=[6,13,7,6,8,8,8,4,7,5,5,6,8,6,6,6,10]
'''frame3 =Frame(bg="#C0C0C0",width=1600,height=400,bd=5,relief=GROOVE)
frame3.grid_propagate(0)
frame3.grid(row = 2,column = 0)
sb = Scrollbar(frame3)
sb.grid(row = 2,column = 23)'''
frame3 =Frame(bg="#C0C0C0",width=1650,height=400,bd=5,relief=GROOVE)
frame3.grid_propagate(0)
frame3.grid(row = 2,column = 0)
for i in range(0,17):
    lab=Label(frame3,text=WH_str[i],width=ttd[i],height=1,bg="#0000A0",fg="#00FFFF",font=('微軟正黑體',11),relief=GROOVE)
    taba.append(lab)
    lab.grid(row = 0,column = i,padx=1,pady=2)
lab=Label(frame3,text=WH_str[17],width=15,height=1,bg="#0000A0",fg="#00FFFF",font=('微軟正黑體',11),relief=GROOVE)
lab.grid(row = 0,column = 17,padx=1,pady=2,columnspan=2)
taba.append(lab)
lab=Label(frame3,text=WH_str[18],width=12,height=1,bg="#0000A0",fg="#00FFFF",font=('微軟正黑體',11),relief=GROOVE)
lab.grid(row = 0,column = 19,padx=1,pady=2,columnspan=2)
taba.append(lab)
for i in range(1,13):
    s=list()
    for j in range(0,19):
        if(j<17):
            s.append(StringVar())
            s[j].set("")
        else:
            s.append(IntVar())
            s[j].set(0)
    for k in range(0,14):
        if(k==3):
            lab=Label(frame3,textvariable=s[k],width=ttd[k],height=1,bg="#FAF0E6",fg="#FF3333",font=('微軟正黑體',11,'bold'),relief=GROOVE)
        elif(k==10 or k==9):
            lab=Label(frame3,textvariable=s[k],width=ttd[k],height=1,bg="#FAF0E6",fg="#006400",font=('微軟正黑體',11,'bold'),relief=GROOVE)
        elif(k==8):
            lab=Label(frame3,textvariable=s[k],width=ttd[k],height=1,bg="#FAF0E6",fg="#CD5555",font=('微軟正黑體',11,'bold'),relief=GROOVE)
        elif(k==11):
            lab=Label(frame3,textvariable=s[k],width=ttd[k],height=1,bg="#FAF0E6",fg="#191970",font=('微軟正黑體',11,'bold'),relief=GROOVE)
        elif(k==2):
            lab=Label(frame3,textvariable=s[k],width=ttd[k],height=1,bg="#FAF0E6",fg="#000000",font=('微軟正黑體',11,'bold'),relief=GROOVE)
        else:
            lab=Label(frame3,textvariable=s[k],width=ttd[k],height=1,bg="#FAF0E6",fg="#292421",font=('微軟正黑體',11),relief=GROOVE)
        lab.grid(row = i,column = k,padx=1,pady=2)
        taba.append(lab)
    entry = Entry(frame3,textvariable = s[14],width=6,font=('微軟正黑體',11),relief=GROOVE)
    entry.grid(row = i,column = 14,padx=1,pady=2)
    taba.append(entry)
    entry = Entry(frame3,textvariable = s[15],width=6,font=('微軟正黑體',11),relief=GROOVE)
    entry.grid(row = i,column = 15,padx=1,pady=2)
    taba.append(entry)
    entry = Entry(frame3,textvariable = s[16],width=10,font=('微軟正黑體',11),relief=GROOVE)
    entry.grid(row = i,column = 16,padx=1,pady=2)
    taba.append(entry)
    R=Radiobutton(frame3,variable = s[17] ,text ="0.07125%",value = 0,command=lambda k=s[17],kd=i-1,a=0,ff=0: hi(a,k,kd,ff),width=6,height=1,bg="#FFFFCD",fg="#292421",font=('微軟正黑體',8),relief=GROOVE)
    R.grid(row=i,column=17,padx=1,pady=0)
    taba.append(R)
    R=Radiobutton(frame3,variable = s[17] ,text ="0.1425%",value = 1,command=lambda k=s[17],kd=i-1,a=1,ff=0: hi(a,k,kd,ff),width=6,height=1,bg="#FFFFCD",fg="#292421",font=('微軟正黑體',8),relief=GROOVE)
    R.grid(row=i,column=18,padx=1,pady=0)
    taba.append(R)
    R=Radiobutton(frame3,variable = s[18] ,text ="0.3%",value = 0,command=lambda k=s[18],kd=i-1,a=0,ff=1: hi(a,k,kd,ff),width=4,height=1,bg="#FFFFCD",fg="#292421",font=('微軟正黑體',8),relief=GROOVE)
    R.grid(row=i,column=19,padx=1,pady=0)
    taba.append(R)
    R=Radiobutton(frame3,variable = s[18] ,text ="0.1%",value = 1,command=lambda k=s[18],kd=i-1,a=1,ff=1: hi(a,k,kd,ff),width=4,height=1,bg="#FFFFCD",fg="#292421",font=('微軟正黑體',8),relief=GROOVE)
    R.grid(row=i,column=20,padx=1,pady=0)
    taba.append(R)
    k=Button(frame3,text= "賣出匯入",command=lambda kd=i-1: sell(kd),width=10, height=1,bg="#004040",fg="#FFFFFF",font=('微軟正黑體',8),relief=GROOVE)
    k.grid(row = i,column = 21,padx=1,pady=0)
    taba.append(k)
    k=Button(frame3,text= "明細",command=lambda kd=i-1: gonext(kd),width=4, height=1,bg="#004040",fg="#FFFFFF",font=('微軟正黑體',8),relief=GROOVE)
    k.grid(row = i,column = 22,padx=1,pady=0)
    taba.append(k)
    
    k=Button(frame3,text= "上",command=lambda kd=i-1: up(kd),width=2, height=1,bg="#004040",fg="#FFFFFF",font=('微軟正黑體',8),relief=GROOVE)
    k.grid(row = i,column = 23,padx=1,pady=0)
    taba.append(k)
    
    k=Button(frame3,text= "下",command=lambda kd=i-1: down(kd),width=2, height=1,bg="#004040",fg="#FFFFFF",font=('微軟正黑體',8),relief=GROOVE)
    k.grid(row = i,column = 24,padx=1,pady=0)
    taba.append(k)
    
    table.append(s)
    del s
#--------------------------------------------------------------------
#------------------------------------------------------交易明細------------------------
it_str=["證券編號","證券名稱","交易類別","交易日期","數量","成交價","手續費","交易稅","當沖","現金股息","總價"]
tta=[8,18,12,12,12,12,12,12,12,12,12]
for i in range(0,11):
    lab=Label(frame3,text=it_str[i],width=tta[i],height=1,bg="#0000A0",fg="#00FFFF",font=('微軟正黑體',12),relief=GROOVE)
    tabb.append(lab)
    lab.grid(row = 0,column = i,padx=2,pady=2)
for i in range(1,13):
    s=list()
    for j in range(0,11):
        s.append(StringVar())
        s[j].set("")
        lab=Label(frame3,textvariable=s[j],width=tta[j],height=1,bg="#FAF0E6",fg="#292421",font=('微軟正黑體',12),relief=GROOVE)
        lab.grid(row = i,column = j,padx=2,pady=2)
        tabb.append(lab)
    k=Button(frame3,text= "刪除項目",command=lambda kd=i-1: delitem(kd),width=8, height=1,bg="#004040",fg="#FFFFFF",font=('微軟正黑體',8),relief=GROOVE)
    k.grid(row = i,column = 11,padx=1,pady=0)
    tabb.append(k)
    gtable.append(s)
    del s
#----------------------------
for i in range(0,len(tabb)):
    tabb[i].grid_remove()
#------------------------------溝通表---------------------------
frame4 =Frame(bg="#C0C0C0",width=1650,height=50,relief=GROOVE)
frame4.grid_propagate(0)
frame4.grid(row = 3,column = 0)
k=Button(frame4,text= "上一頁",command=lambda : pageup(),width=10, height=1,bg="#004040",fg="#FFFFFF",font=('微軟正黑體',16),relief=GROOVE)
k.grid(row = 0,column = 0,padx=2,pady=2)

pa=StringVar()
pa.set(str(page)+"/"+str(int((len(data)-1)/10)+1))
lab=Label(frame4,textvariable=pa,width=10,height=1,bg="#FFFFFF",fg="#000000",font=('微軟正黑體',21),relief=GROOVE)
lab.grid(row = 0,column = 1,padx=2,pady=2)

k=Button(frame4,text= "下一頁",command=lambda : pagedown(),width=10, height=1,bg="#004040",fg="#FFFFFF",font=('微軟正黑體',16),relief=GROOVE)
k.grid(row = 0,column = 2,padx=2,pady=2)


lab=Label(frame4,text="即時對話框：",width=12,height=1,bg="#FFD700",fg="#000000",font=('微軟正黑體',21),relief=GROOVE)
lab.grid(row = 0,column = 3,padx=2,pady=2)

ha=StringVar()
ha.set("歡迎使用本軟體，目前開啟的檔案為："+file)
lab=Label(frame4,textvariable=ha,width=56,height=1,bg="#FFFFFF",fg="#000000",font=('微軟正黑體',21),anchor=W,relief=GROOVE)
lab.grid(row = 0,column = 4,padx=2,pady=2)
#-------------------------------主功能表----------------------------
total=["即時市值","投資成本","賣出匯入","當沖","現金股利","手續費","交易稅","帳面損益","損益評估"]
bcolor=["#8A360F","#FAFFF0","#FF6100","#33AC19","#601080","#03A891","#4040A0","#FF4500","#B22222"]
fcolor=["#FFFFCD","#FF6100","#FAFFF0","#5020A0","#23E8A1","#601080","#00FF00","#B0FFB0","#A0FFA0"]
frame5 =Frame(bg="#C0C0C0",width=1650,height=125,bd=10,relief=GROOVE)
frame5.grid_propagate(0)
frame5.grid(row = 4,column = 0)
dta=list()
tabc=list()
for i in range(0,9):
    dta.append(StringVar())
    dta[i].set(total[i])
for i in range(0,9):
    lab=Label(frame5,textvariable=dta[i],width=13,height=1,bg=bcolor[i],fg=fcolor[i],font=('微軟正黑體',16),relief=GROOVE)
    lab.grid(row = 0,column = i,padx=3,pady=2)
for i in range(0,9):
    dtable.append(StringVar())
for i in range(0,9):
    if(i==5 or i==6):
        lab=Label(frame5,textvariable=dtable[i],width=13,height=2,bg="#FAF0E6",fg="#006400",font=('微軟正黑體',16,'bold'),relief=GROOVE)
    elif(i==4):
        lab=Label(frame5,textvariable=dtable[i],width=13,height=2,bg="#FAF0E6",fg="#CD5555",font=('微軟正黑體',16,'bold'),relief=GROOVE)
    else:
        lab=Label(frame5,textvariable=dtable[i],width=13,height=2,bg="#FAF0E6",fg="#292421",font=('微軟正黑體',16),relief=GROOVE)
    lab.grid(row = 1,column = i,padx=3,pady=2)
    tabc.append(lab)
#------------------------------------------------------------------------------------------
toy=["賣出匯入","賣出淨利","平均成交價","買入股數","買入成本","手續費","現金股利","股票股利","買賣報酬率"]
#-------------------------更新
frame6 =Frame(bg="#C0C0C0",width=1650,height=50,relief=GROOVE)
frame6.grid_propagate(0)
frame6.grid(row = 5,column = 0)
ga=StringVar()
ga.set("")
lab=Label(frame6,textvariable=ga,width=19,height=1,bg="#FFFFFF",fg="#000000",font=('微軟正黑體',18),anchor=W,relief=GROOVE)
lab.grid(row = 0,column = 0,padx=1,pady=2)
lab=Label(frame6,text= "檔案名稱",width=8,height=1,bg="#AFEEEE",fg="#000000",font=('微軟正黑體',18),relief=GROOVE)
lab.grid(row = 0,column = 1,padx=1,pady=2)
ggg=StringVar()
ggg.set(file)
lab=Label(frame6,textvariable=ggg,width=18,height=1,bg="#FFFFFF",fg="#000000",font=('微軟正黑體',18),relief=GROOVE)
lab.grid(row = 0,column = 2,padx=1,pady=2)

lab=Label(frame6,text= "選擇檔案",width=8, height=1,bg="#AFEEEE",fg="#000000",font=('微軟正黑體',18),relief=GROOVE)
lab.grid(row = 0,column = 3,padx=1,pady=2)
global combo
combo=ttk.Combobox(frame6,value=myfile,width=18,font=('微軟正黑體',18))#,state="readonly"
combo.grid(row = 0,column = 4,padx=2,pady=2)
combo.bind("<<ComboboxSelected>>",callbackFunc)
k=Button(frame6,text= "新增檔案",command=renewf,width=8, height=1,bg="#004040",fg="#FFFFFF",font=('微軟正黑體',16),relief=GROOVE)
k.grid(row = 0,column = 5,padx=2,pady=2)
k=Button(frame6,text= "另存新檔",command=newf,width=8, height=1,bg="#004040",fg="#FFFFFF",font=('微軟正黑體',16),relief=GROOVE)
k.grid(row = 0,column = 6,padx=2,pady=2)
k=Button(frame6,text= "開啟檔案",command=openf,width=8, height=1,bg="#004040",fg="#FFFFFF",font=('微軟正黑體',16),relief=GROOVE)
k.grid(row = 0,column = 7,padx=2,pady=2)
k=Button(frame6,text= "更新資料",command=updata,width=8, height=1,bg="#004040",fg="#FFFFFF",font=('微軟正黑體',16),relief=GROOVE)
k.grid(row = 0,column = 8,padx=2,pady=2)
k=Button(frame6,text= "退出程式",command=ex,width=8, height=1,bg="#004040",fg="#FFFFFF",font=('微軟正黑體',16),relief=GROOVE)
k.grid(row = 0,column = 9,padx=2,pady=2)
#---------------------------------------
for i in range(0,12):
    put(i)
    
root.mainloop()#執行





















